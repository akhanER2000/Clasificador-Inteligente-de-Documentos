"""
Generador de archivos Excel (.xlsx) a partir de datos estructurados.
Usa openpyxl para crear hojas de cálculo formateadas.
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel(extracted_data: dict, document_type: str, output_dir: str = ".tmp") -> str:
    """
    Genera un archivo Excel a partir de datos extraídos.
    
    Args:
        extracted_data: Dict con tables y metadata del documento.
        document_type: Tipo de documento clasificado.
        output_dir: Directorio de salida.
        
    Returns:
        Ruta al archivo generado.
    """
    os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{document_type}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb = Workbook()
    
    # Estilos
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    title_font = Font(bold=True, size=14, color="2C3E50")
    metadata_font = Font(italic=True, size=10, color="555555")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    tables = extracted_data.get("tables", [])
    metadata = extracted_data.get("metadata", {})
    title = extracted_data.get("title", "Documento")
    
    if not tables:
        # Si no hay tablas, crear una hoja con metadata
        ws = wb.active
        ws.title = "Datos"
        ws["A1"] = title
        ws["A1"].font = title_font
        
        row = 3
        for key, value in metadata.items():
            ws.cell(row=row, column=1, value=key.capitalize())
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))
            row += 1
    else:
        # Crear una hoja por cada tabla
        for i, table in enumerate(tables):
            if i == 0:
                ws = wb.active
            else:
                ws = wb.create_sheet()
            
            sheet_name = table.get("sheet_name", f"Hoja{i+1}")
            # Limitar nombre de hoja a 31 caracteres (límite de Excel)
            ws.title = sheet_name[:31]
            
            # Título del documento
            ws["A1"] = title
            ws["A1"].font = title_font
            
            # Metadata en fila 2
            if metadata:
                col = 1
                for key, value in metadata.items():
                    if value:
                        ws.cell(row=2, column=col, value=f"{key}: {value}")
                        ws.cell(row=2, column=col).font = metadata_font
                        col += 1
            
            # Headers de la tabla (fila 4)
            headers = table.get("headers", [])
            start_row = 4
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            
            # Datos
            rows = table.get("rows", [])
            for row_idx, row_data in enumerate(rows, start_row + 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
            
            # Ajustar ancho de columnas
            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                max_length = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 10
                
                for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(rows), min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                
                ws.column_dimensions[col_letter].width = min(max_length + 4, 50)
    
    wb.save(filepath)
    return filepath
